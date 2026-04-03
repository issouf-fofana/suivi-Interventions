# ============================================================
# main.py - Backend FastAPI pour le Suivi des Interventions
# ============================================================

import sqlite3
import csv
import io
import json
import hashlib
import os
import shutil
import secrets
import time
from datetime import datetime, date, timedelta
from typing import Optional, List
from fastapi import FastAPI, HTTPException, Query, UploadFile, File, Request, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse, JSONResponse, PlainTextResponse, FileResponse
from pydantic import BaseModel

# ---- Dossier pièces jointes ----
UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

# --- Configuration ---
app = FastAPI(title="Suivi des Interventions", version="3.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:8000",
        "http://127.0.0.1:8000",
        "http://10.0.80.107:8000",
    ],
    allow_origin_regex=r"https?://10\.0\.80\.107(?::\d+)?",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

DB_PATH = "interventions.db"
CONFIG_PATH = "config.json"

# Servir les fichiers statiques (CSS, JS, images...)
_STATIC_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static")
app.mount("/static", StaticFiles(directory=_STATIC_DIR), name="static")

_UPLOAD_STATIC_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")
os.makedirs(_UPLOAD_STATIC_DIR, exist_ok=True)


# ============================================================
# RATE LIMITING (anti brute-force login)
# ============================================================

_login_attempts: dict = {}  # {ip: [timestamp, ...]}
RATE_LIMIT_MAX = 10          # max tentatives
RATE_LIMIT_WINDOW = 300      # fenêtre en secondes (5 min)

def check_rate_limit(ip: str):
    now = time.time()
    attempts = [t for t in _login_attempts.get(ip, []) if now - t < RATE_LIMIT_WINDOW]
    if len(attempts) >= RATE_LIMIT_MAX:
        raise HTTPException(status_code=429, detail="Trop de tentatives. Réessayez dans 5 minutes.")
    attempts.append(now)
    _login_attempts[ip] = attempts

def reset_rate_limit(ip: str):
    _login_attempts.pop(ip, None)


# ============================================================
# AUTH HELPERS
# ============================================================

def hash_password(password: str) -> str:
    """Retourne le sha256 du mot de passe."""
    return hashlib.sha256(password.encode()).hexdigest()

def verifier_token(token: str) -> Optional[dict]:
    """
    Vérifie le token.
    Priorité 1 : session token dans la table `sessions` (avec expiry).
    Priorité 2 : token legacy sha256(username:password_hash) pour rétro-compat.
    """
    if not token:
        return None
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    now_iso = datetime.utcnow().isoformat()

    # Chercher dans sessions
    try:
        sess = conn.execute(
            "SELECT s.*, u.* FROM sessions s JOIN users u ON s.user_id=u.id "
            "WHERE s.token=? AND s.expires_at > ? AND u.actif=1",
            (token, now_iso)
        ).fetchone()
        if sess:
            conn.close()
            return dict(sess)
    except Exception:
        pass

    # Fallback token legacy
    users = conn.execute("SELECT * FROM users WHERE actif=1").fetchall()
    conn.close()
    for u in users:
        expected = hashlib.sha256(f"{u['username']}:{u['password_hash']}".encode()).hexdigest()
        if token == expected:
            return dict(u)
    return None

def get_current_user(request: Request) -> Optional[dict]:
    token = (
        request.headers.get("X-Auth-Token")
        or request.cookies.get("auth_token")
        or request.query_params.get("token")
    )
    return verifier_token(token)

def require_admin(request: Request):
    user = get_current_user(request)
    if not user or user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    return user


def require_admin_or_manager(request: Request):
    user = get_current_user(request)
    if not user or user.get("role") not in ("admin", "manager"):
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs et managers")
    return user


# ============================================================
# MIDDLEWARE AUTH
# ============================================================

@app.middleware("http")
async def middleware_auth(request: Request, call_next):
    path = request.url.path

    routes_publiques = ["/api/auth/login"]
    if request.method == "OPTIONS":
        return await call_next(request)

    if path in routes_publiques or not path.startswith("/api/"):
        return await call_next(request)

    token = request.headers.get("X-Auth-Token") or request.cookies.get("auth_token")

    if not verifier_token(token):
        return JSONResponse(
            status_code=401,
            content={"detail": "Non authentifié. Veuillez vous connecter."}
        )

    return await call_next(request)


# ============================================================
# INIT DB
# ============================================================

def init_db():
    conn = sqlite3.connect(DB_PATH)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS interventions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            prestataire TEXT NOT NULL,
            type_intervention TEXT NOT NULL,
            mois TEXT,
            annee INTEGER,
            date_debut TEXT,
            heure_debut TEXT,
            date_fin TEXT,
            heure_fin TEXT,
            duree_minutes INTEGER,
            site TEXT,
            travaux TEXT,
            prochaine_intervention TEXT,
            notes TEXT,
            statut TEXT DEFAULT 'En cours',
            technicien TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    # Migrations douces
    for col, coldef in [
        ("notes", "TEXT"),
        ("statut", "TEXT DEFAULT 'En cours'"),
        ("technicien", "TEXT"),
    ]:
        try:
            conn.execute(f"ALTER TABLE interventions ADD COLUMN {col} {coldef}")
        except Exception:
            pass

    # Table historique des modifications
    conn.execute("""
        CREATE TABLE IF NOT EXISTS historique (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            intervention_id INTEGER NOT NULL,
            action TEXT NOT NULL,
            champ_modifie TEXT,
            ancienne_valeur TEXT,
            nouvelle_valeur TEXT,
            user_id INTEGER,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    try:
        conn.execute("ALTER TABLE historique ADD COLUMN user_id INTEGER")
    except Exception:
        pass

    # Table users
    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'user',
            actif INTEGER NOT NULL DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Table sessions (tokens avec expiration)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS sessions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            token TEXT NOT NULL UNIQUE,
            user_id INTEGER NOT NULL,
            expires_at TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
        )
    """)

    # Table commentaires
    conn.execute("""
        CREATE TABLE IF NOT EXISTS commentaires (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            intervention_id INTEGER NOT NULL,
            user_id INTEGER,
            username TEXT,
            contenu TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(intervention_id) REFERENCES interventions(id) ON DELETE CASCADE
        )
    """)

    # Table pièces jointes
    conn.execute("""
        CREATE TABLE IF NOT EXISTS pieces_jointes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            intervention_id INTEGER NOT NULL,
            user_id INTEGER,
            nom_original TEXT NOT NULL,
            nom_stockage TEXT NOT NULL,
            taille INTEGER,
            type_mime TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(intervention_id) REFERENCES interventions(id) ON DELETE CASCADE
        )
    """)

    # Table audit_logs
    conn.execute("""
        CREATE TABLE IF NOT EXISTS audit_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            username TEXT,
            action TEXT NOT NULL,
            ressource TEXT,
            ressource_id INTEGER,
            detail TEXT,
            ip TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Compte admin par défaut
    existing_admin = conn.execute("SELECT id FROM users WHERE role='admin' LIMIT 1").fetchone()
    if not existing_admin:
        admin_hash = hashlib.sha256("admin123".encode()).hexdigest()
        conn.execute(
            "INSERT INTO users (username, password_hash, role, actif) VALUES (?, ?, 'admin', 1)",
            ("admin", admin_hash)
        )

    conn.commit()
    conn.close()


init_db()

# ============================================================
# ROUTES PAGES (URLs propres sans .html)
# ============================================================

STATIC_DIR = os.path.join(os.path.dirname(__file__), "static")

def serve(filename: str):
    return FileResponse(os.path.join(STATIC_DIR, filename), media_type="text/html")

@app.get("/", include_in_schema=False)
async def page_dashboard():
    return serve("index.html")

@app.get("/calendrier", include_in_schema=False)
async def page_calendrier():
    return serve("calendar.html")

@app.get("/nouvelle-intervention", include_in_schema=False)
async def page_nouvelle_intervention():
    return serve("form.html")

@app.get("/detail", include_in_schema=False)
async def page_detail():
    return serve("detail.html")

@app.get("/admin", include_in_schema=False)
async def page_admin():
    return serve("admin.html")

@app.get("/connexion", include_in_schema=False)
async def page_connexion():
    return serve("login.html")

# Aliases .html directs (pour les liens relatifs entre pages)
@app.get("/index.html", include_in_schema=False)
async def alias_index(): return serve("index.html")

@app.get("/calendar.html", include_in_schema=False)
async def alias_calendar(): return serve("calendar.html")

@app.get("/form.html", include_in_schema=False)
async def alias_form(): return serve("form.html")

@app.get("/detail.html", include_in_schema=False)
async def alias_detail(): return serve("detail.html")

@app.get("/admin.html", include_in_schema=False)
async def alias_admin(): return serve("admin.html")

@app.get("/login.html", include_in_schema=False)
async def alias_login(): return serve("login.html")

@app.get("/404.html", include_in_schema=False)
async def alias_404(): return serve("404.html")

# ============================================================
# METRICS (minimal placeholder)
# ============================================================
@app.get("/metrics", response_class=PlainTextResponse)
async def metrics():
    """
    Endpoint pour les sondes (ex: Kubernetes liveness/readiness).
    Retourne une ligne de texte simple pour éviter les 404 dans les logs.
    """
    return "service_up 1\n"


# ============================================================
# MODÈLES
# ============================================================

class InterventionCreate(BaseModel):
    prestataire: str
    type_intervention: str
    mois: Optional[str] = None
    annee: Optional[int] = None
    date_debut: Optional[str] = None
    heure_debut: Optional[str] = None
    date_fin: Optional[str] = None
    heure_fin: Optional[str] = None
    duree_minutes: Optional[int] = None
    site: Optional[str] = None
    travaux: Optional[str] = None
    prochaine_intervention: Optional[str] = None
    notes: Optional[str] = None
    statut: Optional[str] = "En cours"
    technicien: Optional[str] = None


class InterventionUpdate(InterventionCreate):
    pass


class CommentaireCreate(BaseModel):
    contenu: str


class AuditAction(BaseModel):
    action: str
    detail: Optional[str] = None


class LoginRequest(BaseModel):
    username: str
    password: str

class UserCreate(BaseModel):
    username: str
    password: str
    role: str = "user"

class UserUpdate(BaseModel):
    username: Optional[str] = None
    password: Optional[str] = None
    role: Optional[str] = None
    actif: Optional[int] = None

class ChangePasswordRequest(BaseModel):
    ancien_mdp: str
    nouveau_mdp: str


# ============================================================
# HELPERS
# ============================================================

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def row_to_dict(row):
    return dict(row)

def calcul_duree(date_debut, heure_debut, date_fin, heure_fin):
    try:
        if date_debut and heure_debut and date_fin and heure_fin:
            dt1 = datetime.strptime(f"{date_debut} {heure_debut}", "%Y-%m-%d %H:%M")
            dt2 = datetime.strptime(f"{date_fin} {heure_fin}", "%Y-%m-%d %H:%M")
            delta = dt2 - dt1
            return max(0, int(delta.total_seconds() / 60))
    except Exception:
        pass
    return None

def deduire_mois_annee(date_debut, mois, annee):
    if date_debut and not mois:
        try:
            d = datetime.strptime(date_debut, "%Y-%m-%d")
            noms = ["janvier","février","mars","avril","mai","juin",
                    "juillet","août","septembre","octobre","novembre","décembre"]
            return noms[d.month - 1], d.year
        except Exception:
            pass
    return mois, annee

def build_where(prestataire, type_intervention, mois, annee, site, search,
                annee_val=None, date_debut_from=None, date_debut_to=None):
    """Construit la clause WHERE dynamique."""
    where = "WHERE 1=1"
    params = []
    if prestataire:
        where += " AND LOWER(prestataire) = LOWER(?)"
        params.append(prestataire)
    if type_intervention:
        where += " AND LOWER(type_intervention) = LOWER(?)"
        params.append(type_intervention)
    if mois:
        where += " AND LOWER(mois) = LOWER(?)"
        params.append(mois)
    if annee:
        where += " AND annee = ?"
        params.append(annee)
    if annee_val:
        where += " AND annee = ?"
        params.append(annee_val)
    if date_debut_from:
        where += " AND date_debut >= ?"
        params.append(date_debut_from)
    if date_debut_to:
        where += " AND date_debut <= ?"
        params.append(date_debut_to)
    if site:
        where += " AND LOWER(site) LIKE LOWER(?)"
        params.append(f"%{site}%")
    if search:
        where += """ AND (
            LOWER(prestataire) LIKE LOWER(?) OR
            LOWER(site) LIKE LOWER(?) OR
            LOWER(travaux) LIKE LOWER(?) OR
            LOWER(type_intervention) LIKE LOWER(?)
        )"""
        s = f"%{search}%"
        params.extend([s, s, s, s])
    return where, params


def enregistrer_historique(conn, intervention_id: int, action: str,
                            champ=None, ancienne=None, nouvelle=None, user_id=None):
    """Insère une entrée dans la table historique."""
    conn.execute("""
        INSERT INTO historique (intervention_id, action, champ_modifie, ancienne_valeur, nouvelle_valeur, user_id)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (intervention_id, action, champ, ancienne, nouvelle, user_id))


def audit(conn, user: Optional[dict], action: str, ressource: str = None,
          ressource_id: int = None, detail: str = None, ip: str = None):
    """Enregistre une action dans le journal d'audit."""
    conn.execute("""
        INSERT INTO audit_logs (user_id, username, action, ressource, ressource_id, detail, ip)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (
        user["id"] if user else None,
        user["username"] if user else "system",
        action, ressource, ressource_id, detail, ip
    ))


# ============================================================
# ENDPOINTS AUTH
# ============================================================

@app.post("/api/auth/login")
def login(data: LoginRequest, request: Request):
    """Authentifie l'utilisateur avec username + password."""
    ip = request.client.host if request.client else "unknown"
    check_rate_limit(ip)

    conn = get_db()
    user = conn.execute(
        "SELECT * FROM users WHERE username=? AND actif=1", (data.username,)
    ).fetchone()
    if not user or hash_password(data.password) != user["password_hash"]:
        conn.close()
        raise HTTPException(status_code=401, detail="Identifiants incorrects")

    reset_rate_limit(ip)

    # Créer une session avec expiration 8h
    token = secrets.token_hex(32)
    expires_at = (datetime.utcnow() + timedelta(hours=8)).isoformat()
    conn.execute(
        "INSERT INTO sessions (token, user_id, expires_at) VALUES (?, ?, ?)",
        (token, user["id"], expires_at)
    )
    # Nettoyer les vieilles sessions de cet utilisateur (garder les 5 dernières)
    conn.execute("""
        DELETE FROM sessions WHERE user_id=? AND id NOT IN (
            SELECT id FROM sessions WHERE user_id=? ORDER BY created_at DESC LIMIT 5
        )
    """, (user["id"], user["id"]))

    audit(conn, dict(user), "login", ip=ip)
    conn.commit()
    conn.close()

    return {
        "token": token,
        "username": user["username"],
        "role": user["role"],
        "expires_at": expires_at,
        "message": "Connexion réussie"
    }


@app.post("/api/auth/logout")
def logout(request: Request):
    token = request.headers.get("X-Auth-Token") or request.cookies.get("auth_token")
    if token:
        conn = get_db()
        user = get_current_user(request)
        conn.execute("DELETE FROM sessions WHERE token=?", (token,))
        if user:
            audit(conn, user, "logout")
        conn.commit()
        conn.close()
    return {"message": "Déconnecté"}


@app.get("/api/auth/me")
def get_me(request: Request):
    """Retourne les infos de l'utilisateur connecté."""
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Non authentifié")
    return {"id": user["id"], "username": user["username"], "role": user["role"]}


@app.post("/api/auth/change-password")
def change_password(data: ChangePasswordRequest, request: Request):
    """Change le mot de passe de l'utilisateur connecté."""
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Non authentifié")
    if hash_password(data.ancien_mdp) != user["password_hash"]:
        raise HTTPException(status_code=400, detail="Ancien mot de passe incorrect")
    if len(data.nouveau_mdp) < 4:
        raise HTTPException(status_code=400, detail="Le mot de passe doit faire au moins 4 caractères")
    new_hash = hash_password(data.nouveau_mdp)
    conn = get_db()
    conn.execute("UPDATE users SET password_hash=? WHERE id=?", (new_hash, user["id"]))
    conn.commit()
    conn.close()
    # Retourner le nouveau token
    new_token = hashlib.sha256(f"{user['username']}:{new_hash}".encode()).hexdigest()
    return {"message": "Mot de passe modifié", "token": new_token}


# ============================================================
# ENDPOINTS USERS (admin seulement)
# ============================================================

@app.get("/api/users")
def liste_users(request: Request):
    require_admin_or_manager(request)
    conn = get_db()
    rows = conn.execute("SELECT id, username, role, actif, created_at FROM users ORDER BY id").fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/api/users", status_code=201)
def creer_user(data: UserCreate, request: Request):
    user = require_admin_or_manager(request)
    if data.role not in ("admin", "manager", "user"):
        raise HTTPException(status_code=400, detail="Rôle invalide (admin, manager ou user)")
    if user["role"] == "manager" and data.role == "admin":
        raise HTTPException(status_code=403, detail="Un manager ne peut pas créer un admin")
    if len(data.password) < 4:
        raise HTTPException(status_code=400, detail="Mot de passe trop court (min 4 caractères)")
    conn = get_db()
    existing = conn.execute("SELECT id FROM users WHERE username=?", (data.username,)).fetchone()
    if existing:
        conn.close()
        raise HTTPException(status_code=409, detail="Ce nom d'utilisateur existe déjà")
    pw_hash = hash_password(data.password)
    cur = conn.execute(
        "INSERT INTO users (username, password_hash, role, actif) VALUES (?,?,?,1)",
        (data.username, pw_hash, data.role)
    )
    conn.commit()
    user_id = cur.lastrowid
    audit(conn, user, "create_user", "user", user_id, detail=f"role={data.role}")
    conn.close()
    return {"id": user_id, "username": data.username, "role": data.role, "actif": 1}


@app.put("/api/users/{user_id}")
def modifier_user(user_id: int, data: UserUpdate, request: Request):
    current_admin = require_admin_or_manager(request)
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if not user:
        conn.close()
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")

    if current_admin["role"] == "manager":
        if data.role == "admin":
            conn.close()
            raise HTTPException(status_code=403, detail="Un manager ne peut pas promouvoir un utilisateur en admin")
        if data.actif == 0:
            conn.close()
            raise HTTPException(status_code=403, detail="Un manager ne peut pas désactiver un utilisateur")

    if data.actif == 0 and current_admin["id"] == user_id:
        conn.close()
        raise HTTPException(status_code=400, detail="Vous ne pouvez pas désactiver votre propre compte")

    username = data.username or user["username"]
    role = data.role if data.role in ("admin", "manager", "user") else user["role"]
    actif = data.actif if data.actif is not None else user["actif"]
    pw_hash = hash_password(data.password) if data.password else user["password_hash"]

    if data.password and len(data.password) < 4:
        conn.close()
        raise HTTPException(status_code=400, detail="Mot de passe trop court (min 4 caractères)")

    # Vérifier unicité username si changé
    if username != user["username"]:
        existing = conn.execute("SELECT id FROM users WHERE username=? AND id!=?", (username, user_id)).fetchone()
        if existing:
            conn.close()
            raise HTTPException(status_code=409, detail="Ce nom d'utilisateur existe déjà")

    conn.execute(
        "UPDATE users SET username=?, password_hash=?, role=?, actif=? WHERE id=?",
        (username, pw_hash, role, actif, user_id)
    )
    conn.commit()
    audit(conn, current_admin, "modify_user", "user", user_id,
          detail=f"username={username},role={role},actif={actif}")
    conn.close()
    new_token = hashlib.sha256(f"{username}:{pw_hash}".encode()).hexdigest()
    return {"id": user_id, "username": username, "role": role, "actif": actif, "new_token": new_token}


@app.delete("/api/users/{user_id}", status_code=204)
def supprimer_user(user_id: int, request: Request):
    current_admin = require_admin(request)
    if current_admin["id"] == user_id:
        raise HTTPException(status_code=400, detail="Vous ne pouvez pas supprimer votre propre compte")
    conn = get_db()
    user = conn.execute("SELECT id, username FROM users WHERE id=?", (user_id,)).fetchone()
    if not user:
        conn.close()
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    conn.execute("DELETE FROM users WHERE id=?", (user_id,))
    conn.commit()
    audit(conn, current_admin, "delete_user", "user", user_id, detail=f"username={user['username']}")
    conn.close()
    return Response(status_code=204)


# ============================================================
# ENDPOINTS CRUD
# ============================================================

@app.get("/api/interventions")
def liste_interventions(
    prestataire: Optional[str] = Query(None),
    type_intervention: Optional[str] = Query(None),
    mois: Optional[str] = Query(None),
    annee: Optional[int] = Query(None),
    site: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    statut: Optional[str] = Query(None),
    technicien: Optional[str] = Query(None),
    date_debut_from: Optional[str] = Query(None),
    date_debut_to: Optional[str] = Query(None),
    page: Optional[int] = Query(None),
    per_page: Optional[int] = Query(None),
    sort: Optional[str] = Query("date_debut"),
    order: Optional[str] = Query("desc"),
):
    conn = get_db()
    where, params = build_where(prestataire, type_intervention, mois, annee, site, search,
                                date_debut_from=date_debut_from, date_debut_to=date_debut_to)
    if statut:
        where += " AND statut = ?"
        params.append(statut)
    if technicien:
        where += " AND LOWER(technicien) LIKE LOWER(?)"
        params.append(f"%{technicien}%")

    # Tri sécurisé
    allowed_sorts = {"date_debut", "id", "prestataire", "site", "duree_minutes", "statut", "technicien", "created_at"}
    sort_col = sort if sort in allowed_sorts else "date_debut"
    sort_dir = "ASC" if order == "asc" else "DESC"

    total = conn.execute(f"SELECT COUNT(*) FROM interventions {where}", params).fetchone()[0]

    if page and per_page:
        offset = (page - 1) * per_page
        rows = conn.execute(
            f"SELECT * FROM interventions {where} ORDER BY {sort_col} {sort_dir}, id DESC LIMIT ? OFFSET ?",
            params + [per_page, offset]
        ).fetchall()
    else:
        rows = conn.execute(
            f"SELECT * FROM interventions {where} ORDER BY {sort_col} {sort_dir}, id DESC",
            params
        ).fetchall()

    conn.close()
    result = [row_to_dict(r) for r in rows]
    if page and per_page:
        return {"data": result, "total": total, "page": page, "per_page": per_page, "pages": -(-total // per_page)}
    return result


@app.post("/api/interventions", status_code=201)
def creer_intervention(data: InterventionCreate, request: Request):
    user = get_current_user(request)
    duree = data.duree_minutes or calcul_duree(data.date_debut, data.heure_debut, data.date_fin, data.heure_fin)
    mois, annee = deduire_mois_annee(data.date_debut, data.mois, data.annee)
    statut = data.statut or "En cours"
    conn = get_db()
    cur = conn.execute("""
        INSERT INTO interventions
        (prestataire, type_intervention, mois, annee, date_debut, heure_debut,
         date_fin, heure_fin, duree_minutes, site, travaux, prochaine_intervention, notes, statut, technicien)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (data.prestataire, data.type_intervention, mois, annee,
          data.date_debut, data.heure_debut, data.date_fin, data.heure_fin,
          duree, data.site, data.travaux, data.prochaine_intervention, data.notes,
          statut, data.technicien))
    new_id = cur.lastrowid

    enregistrer_historique(conn, new_id, "creation", user_id=user["id"] if user else None)
    audit(conn, user, "create_intervention", "intervention", new_id,
          detail=f"{data.prestataire} - {data.site}")

    conn.commit()
    row = conn.execute("SELECT * FROM interventions WHERE id=?", (new_id,)).fetchone()
    conn.close()
    return row_to_dict(row)


@app.get("/api/interventions/{intervention_id}")
def detail_intervention(intervention_id: int):
    conn = get_db()
    row = conn.execute("SELECT * FROM interventions WHERE id=?", (intervention_id,)).fetchone()
    conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="Intervention non trouvée")
    return row_to_dict(row)


@app.put("/api/interventions/{intervention_id}")
def modifier_intervention(intervention_id: int, data: InterventionUpdate, request: Request):
    user = get_current_user(request)
    conn = get_db()
    ancien_row = conn.execute("SELECT * FROM interventions WHERE id=?", (intervention_id,)).fetchone()
    if not ancien_row:
        conn.close()
        raise HTTPException(status_code=404, detail="Intervention non trouvée")

    duree = data.duree_minutes or calcul_duree(data.date_debut, data.heure_debut, data.date_fin, data.heure_fin)
    mois, annee = deduire_mois_annee(data.date_debut, data.mois, data.annee)

    champs_a_comparer = {
        "prestataire": data.prestataire,
        "type_intervention": data.type_intervention,
        "mois": mois, "annee": annee,
        "date_debut": data.date_debut, "heure_debut": data.heure_debut,
        "date_fin": data.date_fin, "heure_fin": data.heure_fin,
        "duree_minutes": duree, "site": data.site,
        "travaux": data.travaux,
        "prochaine_intervention": data.prochaine_intervention,
        "notes": data.notes,
        "statut": data.statut,
        "technicien": data.technicien,
    }
    ancien = row_to_dict(ancien_row)

    conn.execute("""
        UPDATE interventions SET
            prestataire=?, type_intervention=?, mois=?, annee=?,
            date_debut=?, heure_debut=?, date_fin=?, heure_fin=?,
            duree_minutes=?, site=?, travaux=?, prochaine_intervention=?,
            notes=?, statut=?, technicien=?
        WHERE id=?
    """, (data.prestataire, data.type_intervention, mois, annee,
          data.date_debut, data.heure_debut, data.date_fin, data.heure_fin,
          duree, data.site, data.travaux, data.prochaine_intervention,
          data.notes, data.statut, data.technicien, intervention_id))

    uid = user["id"] if user else None
    for champ, nouvelle_val in champs_a_comparer.items():
        ancienne_val = ancien.get(champ)
        if str(ancienne_val or '') != str(nouvelle_val or ''):
            enregistrer_historique(conn, intervention_id, "modification",
                champ=champ,
                ancienne=str(ancienne_val) if ancienne_val is not None else None,
                nouvelle=str(nouvelle_val) if nouvelle_val is not None else None,
                user_id=uid)

    audit(conn, user, "update_intervention", "intervention", intervention_id)
    conn.commit()
    row = conn.execute("SELECT * FROM interventions WHERE id=?", (intervention_id,)).fetchone()
    conn.close()
    return row_to_dict(row)


@app.delete("/api/interventions/{intervention_id}")
def supprimer_intervention(intervention_id: int, request: Request):
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Non authentifié")
    if user.get("role") == "manager":
        raise HTTPException(status_code=403, detail="Les managers ne peuvent pas supprimer des interventions")
    conn = get_db()
    row = conn.execute("SELECT * FROM interventions WHERE id=?", (intervention_id,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Intervention non trouvée")

    snapshot = json.dumps(row_to_dict(row), ensure_ascii=False)
    enregistrer_historique(conn, intervention_id, "suppression",
                           ancienne=snapshot, user_id=user["id"] if user else None)
    audit(conn, user, "delete_intervention", "intervention", intervention_id,
          detail=f"{row['prestataire']} - {row['site']}")

    # Supprimer les fichiers joints physiques
    pjs = conn.execute("SELECT nom_stockage FROM pieces_jointes WHERE intervention_id=?",
                       (intervention_id,)).fetchall()
    for pj in pjs:
        try:
            os.remove(os.path.join(UPLOAD_DIR, pj["nom_stockage"]))
        except Exception:
            pass

    conn.execute("DELETE FROM interventions WHERE id=?", (intervention_id,))
    conn.commit()
    conn.close()
    return {"message": f"Intervention {intervention_id} supprimée"}


# --- Historique d'une intervention ---
@app.get("/api/interventions/{intervention_id}/historique")
def historique_intervention(intervention_id: int):
    conn = get_db()
    rows = conn.execute(
        """SELECT h.*, u.username as auteur FROM historique h
           LEFT JOIN users u ON h.user_id = u.id
           WHERE h.intervention_id = ? ORDER BY h.created_at DESC""",
        (intervention_id,)
    ).fetchall()
    conn.close()
    return [row_to_dict(r) for r in rows]


# ============================================================
# COMMENTAIRES
# ============================================================

@app.get("/api/interventions/{intervention_id}/commentaires")
def liste_commentaires(intervention_id: int):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM commentaires WHERE intervention_id=? ORDER BY created_at ASC",
        (intervention_id,)
    ).fetchall()
    conn.close()
    return [row_to_dict(r) for r in rows]


@app.post("/api/interventions/{intervention_id}/commentaires", status_code=201)
def ajouter_commentaire(intervention_id: int, data: CommentaireCreate, request: Request):
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Non authentifié")
    conn = get_db()
    row = conn.execute("SELECT id FROM interventions WHERE id=?", (intervention_id,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Intervention non trouvée")
    cur = conn.execute(
        "INSERT INTO commentaires (intervention_id, user_id, username, contenu) VALUES (?,?,?,?)",
        (intervention_id, user["id"], user["username"], data.contenu.strip())
    )
    audit(conn, user, "add_comment", "intervention", intervention_id)
    conn.commit()
    new = conn.execute("SELECT * FROM commentaires WHERE id=?", (cur.lastrowid,)).fetchone()
    conn.close()
    return row_to_dict(new)


@app.delete("/api/commentaires/{comment_id}", status_code=204)
def supprimer_commentaire(comment_id: int, request: Request):
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Non authentifié")
    if user.get("role") == "manager":
        raise HTTPException(status_code=403, detail="Les managers ne peuvent pas supprimer des commentaires")
    conn = get_db()
    comm = conn.execute("SELECT * FROM commentaires WHERE id=?", (comment_id,)).fetchone()
    if not comm:
        conn.close()
        raise HTTPException(status_code=404, detail="Commentaire non trouvé")
    # Seul l'auteur ou un admin peut supprimer
    if comm["user_id"] != user["id"] and user.get("role") != "admin":
        conn.close()
        raise HTTPException(status_code=403, detail="Non autorisé")
    conn.execute("DELETE FROM commentaires WHERE id=?", (comment_id,))
    audit(conn, user, "delete_comment", "commentaire", comment_id)
    conn.commit()
    conn.close()
    return Response(status_code=204)


# ============================================================
# PIÈCES JOINTES
# ============================================================

ALLOWED_EXTENSIONS = {".pdf", ".jpg", ".jpeg", ".png", ".gif", ".webp",
                      ".doc", ".docx", ".xls", ".xlsx", ".txt", ".zip"}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB


@app.get("/api/interventions/{intervention_id}/pieces-jointes")
def liste_pieces_jointes(intervention_id: int):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM pieces_jointes WHERE intervention_id=? ORDER BY created_at DESC",
        (intervention_id,)
    ).fetchall()
    conn.close()
    return [row_to_dict(r) for r in rows]


@app.post("/api/interventions/{intervention_id}/pieces-jointes", status_code=201)
async def uploader_piece_jointe(intervention_id: int, fichier: UploadFile = File(...),
                                 request: Request = None):
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Non authentifié")

    conn = get_db()
    row = conn.execute("SELECT id FROM interventions WHERE id=?", (intervention_id,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Intervention non trouvée")

    ext = os.path.splitext(fichier.filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        conn.close()
        raise HTTPException(status_code=400, detail=f"Extension non autorisée. Autorisées : {', '.join(ALLOWED_EXTENSIONS)}")

    content = await fichier.read()
    if len(content) > MAX_FILE_SIZE:
        conn.close()
        raise HTTPException(status_code=400, detail="Fichier trop volumineux (max 10 Mo)")

    nom_stockage = f"{secrets.token_hex(16)}{ext}"
    chemin = os.path.join(UPLOAD_DIR, nom_stockage)
    with open(chemin, "wb") as f:
        f.write(content)

    cur = conn.execute(
        """INSERT INTO pieces_jointes (intervention_id, user_id, nom_original, nom_stockage, taille, type_mime)
           VALUES (?,?,?,?,?,?)""",
        (intervention_id, user["id"], fichier.filename, nom_stockage, len(content), fichier.content_type)
    )
    audit(conn, user, "upload_file", "intervention", intervention_id, detail=fichier.filename)
    conn.commit()
    new = conn.execute("SELECT * FROM pieces_jointes WHERE id=?", (cur.lastrowid,)).fetchone()
    conn.close()
    return row_to_dict(new)


@app.get("/api/pieces-jointes/{pj_id}/telecharger")
def telecharger_piece_jointe(pj_id: int, request: Request):
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Non authentifié")
    conn = get_db()
    pj = conn.execute("SELECT * FROM pieces_jointes WHERE id=?", (pj_id,)).fetchone()
    conn.close()
    if not pj:
        raise HTTPException(status_code=404, detail="Fichier non trouvé")
    chemin = os.path.join(UPLOAD_DIR, pj["nom_stockage"])
    if not os.path.exists(chemin):
        raise HTTPException(status_code=404, detail="Fichier introuvable sur le serveur")
    return FileResponse(chemin, filename=pj["nom_original"], media_type=pj["type_mime"] or "application/octet-stream")


@app.delete("/api/pieces-jointes/{pj_id}", status_code=204)
def supprimer_piece_jointe(pj_id: int, request: Request):
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Non authentifié")
    if user.get("role") == "manager":
        raise HTTPException(status_code=403, detail="Les managers ne peuvent pas supprimer des fichiers")
    conn = get_db()
    pj = conn.execute("SELECT * FROM pieces_jointes WHERE id=?", (pj_id,)).fetchone()
    if not pj:
        conn.close()
        raise HTTPException(status_code=404, detail="Fichier non trouvé")
    if pj["user_id"] != user["id"] and user.get("role") != "admin":
        conn.close()
        raise HTTPException(status_code=403, detail="Non autorisé")
    try:
        os.remove(os.path.join(UPLOAD_DIR, pj["nom_stockage"]))
    except Exception:
        pass
    conn.execute("DELETE FROM pieces_jointes WHERE id=?", (pj_id,))
    audit(conn, user, "delete_file", "piece_jointe", pj_id, detail=pj["nom_original"])
    conn.commit()
    conn.close()
    return Response(status_code=204)


# ============================================================
# AUDIT LOGS
# ============================================================

@app.get("/api/audit-logs")
def liste_audit_logs(
    request: Request,
    limit: int = Query(100, le=500),
    offset: int = Query(0),
    username: Optional[str] = Query(None),
    action: Optional[str] = Query(None),
):
    require_admin_or_manager(request)
    conn = get_db()
    where = "WHERE 1=1"
    params = []
    if username:
        where += " AND username LIKE ?"
        params.append(f"%{username}%")
    if action:
        where += " AND action LIKE ?"
        params.append(f"%{action}%")
    total = conn.execute(f"SELECT COUNT(*) FROM audit_logs {where}", params).fetchone()[0]
    rows = conn.execute(
        f"SELECT * FROM audit_logs {where} ORDER BY created_at DESC LIMIT ? OFFSET ?",
        params + [limit, offset]
    ).fetchall()
    conn.close()
    return {"total": total, "data": [row_to_dict(r) for r in rows]}


# ============================================================
# PDF EXPORT
# ============================================================

@app.get("/api/interventions/{intervention_id}/pdf")
def export_pdf_intervention(intervention_id: int, request: Request):
    """Génère un rapport PDF pour une intervention."""
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Non authentifié")
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        from reportlab.lib.enums import TA_LEFT, TA_CENTER
    except ImportError:
        raise HTTPException(status_code=500, detail="reportlab non installé. Lancez: pip install reportlab")

    conn = get_db()
    row = conn.execute("SELECT * FROM interventions WHERE id=?", (intervention_id,)).fetchone()
    if not row:
        conn.close()
        raise HTTPException(status_code=404, detail="Intervention non trouvée")
    inter = row_to_dict(row)
    comments = [dict(c) for c in conn.execute(
        "SELECT c.*, u.username FROM commentaires c LEFT JOIN users u ON c.user_id=u.id "
        "WHERE c.intervention_id=? ORDER BY c.created_at ASC", (intervention_id,)
    ).fetchall()]
    pjs = [dict(p) for p in conn.execute(
        "SELECT * FROM pieces_jointes WHERE intervention_id=? ORDER BY created_at ASC",
        (intervention_id,)
    ).fetchall()]
    conn.close()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5*cm, bottomMargin=2*cm,
                             leftMargin=1.8*cm, rightMargin=1.8*cm)
    styles = getSampleStyleSheet()
    W = 17.4 * cm  # largeur utile

    DARK_BLUE  = colors.HexColor("#1e2d5a")
    PRIMARY    = colors.HexColor("#3b5bdb")
    LIGHT_GRAY = colors.HexColor("#f1f5f9")
    MID_GRAY   = colors.HexColor("#e2e8f0")
    GREEN      = colors.HexColor("#059669")
    GREEN_BG   = colors.HexColor("#dcfce7")
    RED        = colors.HexColor("#dc2626")
    RED_BG     = colors.HexColor("#fee2e2")
    ORANGE     = colors.HexColor("#d97706")
    ORANGE_BG  = colors.HexColor("#fef9c3")
    BLUE_BG    = colors.HexColor("#dbeafe")
    SLATE      = colors.HexColor("#64748b")
    TEXT       = colors.HexColor("#1e293b")

    def S(name, **kw):
        return ParagraphStyle(name, parent=styles["Normal"], **kw)

    title_s   = S("T", fontSize=22, fontName="Helvetica-Bold", textColor=colors.white)
    sub_s     = S("Su", fontSize=9,  fontName="Helvetica",     textColor=colors.HexColor("#c5d0fc"))
    label_s   = S("L", fontSize=8,  fontName="Helvetica-Bold", textColor=SLATE, spaceAfter=0)
    value_s   = S("V", fontSize=11, fontName="Helvetica-Bold", textColor=TEXT)
    value_sm  = S("Vs", fontSize=10, fontName="Helvetica",     textColor=TEXT, leading=14)
    section_s = S("Sec", fontSize=10, fontName="Helvetica-Bold", textColor=DARK_BLUE)
    body_s    = S("B", fontSize=10, fontName="Helvetica",      textColor=TEXT, leading=15)
    small_s   = S("Sm", fontSize=8, fontName="Helvetica",      textColor=SLATE)
    footer_s  = S("F", fontSize=8, fontName="Helvetica",       textColor=SLATE, alignment=TA_CENTER)

    def fmt_date(d):
        if not d: return "—"
        try:
            p = d.split("-"); return f"{p[2]}/{p[1]}/{p[0]}" if len(p)==3 else d
        except: return d or "—"

    def fmt_duree(dm):
        if not dm: return "—"
        h, m = dm // 60, dm % 60
        return f"{h}h{m:02d}min" if m else f"{h}h"

    dm = inter.get("duree_minutes")
    duree_fmt = fmt_duree(dm)
    statut = inter.get("statut") or "En cours"
    statut_cfg = {
        "Terminée":   (GREEN,  GREEN_BG),
        "En cours":   (PRIMARY, BLUE_BG),
        "En attente": (ORANGE, ORANGE_BG),
        "Annulée":    (RED,    RED_BG),
    }.get(statut, (PRIMARY, BLUE_BG))

    # Jours restants prochaine intervention
    prochaine_str = ""
    if inter.get("prochaine_intervention"):
        prochaine_str = fmt_date(inter["prochaine_intervention"])
        try:
            delta = (datetime.strptime(inter["prochaine_intervention"], "%Y-%m-%d").date()
                     - date.today()).days
            if delta >= 0:
                prochaine_str += f"  (Dans {delta} jour{'s' if delta>1 else ''})"
            else:
                prochaine_str += f"  (Il y a {-delta} jour{'s' if -delta>1 else ''})"
        except: pass

    elements = []

    # ── BANDEAU HEADER ──────────────────────────────────────────
    hdr = Table([[
        Paragraph(f"RAPPORT D'INTERVENTION  N° {inter['id']}", title_s),
        Paragraph(
            f"Généré le {date.today().strftime('%d/%m/%Y')}<br/>"
            f"<font size='8'>par {user['username']}</font>",
            sub_s),
    ]], colWidths=[12*cm, 5.4*cm])
    hdr.setStyle(TableStyle([
        ("BACKGROUND",   (0,0),(-1,-1), DARK_BLUE),
        ("TOPPADDING",   (0,0),(-1,-1), 20),
        ("BOTTOMPADDING",(0,0),(-1,-1), 20),
        ("LEFTPADDING",  (0,0),(0,-1),  20),
        ("RIGHTPADDING", (-1,0),(-1,-1),16),
        ("VALIGN",       (0,0),(-1,-1), "MIDDLE"),
        ("ALIGN",        (1,0),(1,-1),  "RIGHT"),
        ("ROUNDEDCORNERS", [6],),
    ]))
    elements.append(hdr)
    elements.append(Spacer(1, 0.4*cm))

    # ── BLOC PRESTATAIRE + STATUT ────────────────────────────────
    statut_tbl = Table([[
        Paragraph(f"<font color='#64748b' size='8'>STATUT</font><br/>"
                  f"<b>{statut}</b>",
                  S("st", fontSize=11, fontName="Helvetica-Bold", textColor=statut_cfg[0]))
    ]], colWidths=[4*cm])
    statut_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,-1), statut_cfg[1]),
        ("TOPPADDING",    (0,0),(-1,-1), 10),
        ("BOTTOMPADDING", (0,0),(-1,-1), 10),
        ("LEFTPADDING",   (0,0),(-1,-1), 12),
        ("RIGHTPADDING",  (0,0),(-1,-1), 12),
        ("ROUNDEDCORNERS",[6],),
    ]))

    hero = Table([[
        Table([
            [Paragraph('<font color="#64748b" size="8">PRESTATAIRE</font>', label_s)],
            [Paragraph(inter.get("prestataire") or "—", S("prest", fontSize=18, fontName="Helvetica-Bold", textColor=TEXT))],
            [Paragraph(f'{inter.get("site") or "—"}  ·  {inter.get("mois","").capitalize()} {inter.get("annee","") or ""}', small_s)],
        ]),
        statut_tbl
    ]], colWidths=[12*cm, 5.4*cm])
    hero.setStyle(TableStyle([
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("LEFTPADDING",(0,0),(0,-1),0),
        ("RIGHTPADDING",(-1,0),(-1,-1),0),
        ("TOPPADDING",(0,0),(-1,-1),0),
        ("BOTTOMPADDING",(0,0),(-1,-1),0),
    ]))
    elements.append(hero)
    elements.append(Spacer(1, 0.4*cm))

    # ── 4 KPI CARDS ─────────────────────────────────────────────
    def kpi_card(label, val, sub="", col=TEXT):
        inner = Table([[
            Paragraph(f'<font color="#64748b" size="8">{label}</font>', label_s),
            Paragraph(f'<font color="{col.hexval() if hasattr(col,"hexval") else "#1e293b"}">{val}</font>',
                      S("kv", fontSize=14, fontName="Helvetica-Bold", textColor=col)),
            Paragraph(sub, small_s),
        ]], colWidths=[None])
        inner.setStyle(TableStyle([
            ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),
            ("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),10),
        ]))
        outer = Table([[inner]], colWidths=[W/4 - 0.15*cm])
        outer.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1), LIGHT_GRAY),
            ("TOPPADDING",(0,0),(-1,-1),0),("BOTTOMPADDING",(0,0),(-1,-1),0),
            ("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0),
            ("BOX",(0,0),(-1,-1),0.5,MID_GRAY),
        ]))
        return outer

    duree_col = RED if (dm and dm > 480) else TEXT
    kpis = Table([[
        kpi_card("DURÉE",      duree_fmt,                             f"{dm or 0} minutes", duree_col),
        kpi_card("DATE DÉBUT", fmt_date(inter.get("date_debut")),     inter.get("heure_debut") or ""),
        kpi_card("DATE FIN",   fmt_date(inter.get("date_fin")),       inter.get("heure_fin") or ""),
        kpi_card("PROCHAINE",  fmt_date(inter.get("prochaine_intervention")) if inter.get("prochaine_intervention") else "—",
                 prochaine_str.split("(")[-1].rstrip(")") if "(" in prochaine_str else ""),
    ]], colWidths=[W/4]*4)
    kpis.setStyle(TableStyle([
        ("TOPPADDING",(0,0),(-1,-1),0),("BOTTOMPADDING",(0,0),(-1,-1),0),
        ("LEFTPADDING",(0,0),(-1,-1),3),("RIGHTPADDING",(0,0),(-1,-1),3),
    ]))
    elements.append(kpis)
    elements.append(Spacer(1, 0.4*cm))

    # ── TIMELINE ────────────────────────────────────────────────
    dd = f"{fmt_date(inter.get('date_debut'))}  {inter.get('heure_debut') or '—'}"
    df = f"{fmt_date(inter.get('date_fin'))}  {inter.get('heure_fin') or '—'}"
    timeline = Table([[
        Table([[
            [Paragraph("DÉBUT D'INTERVENTION", S("tl", fontSize=8, fontName="Helvetica-Bold", textColor=SLATE))],
            [Paragraph(fmt_date(inter.get("date_debut")), S("td", fontSize=13, fontName="Helvetica-Bold", textColor=TEXT))],
            [Paragraph(inter.get("heure_debut") or "—", S("th", fontSize=12, fontName="Helvetica-Bold", textColor=PRIMARY))],
        ]]),
        Paragraph("→", S("arr", fontSize=16, textColor=SLATE, alignment=1)),
        Table([[
            [Paragraph("DURÉE TOTALE", S("tl2", fontSize=8, fontName="Helvetica-Bold", textColor=SLATE, alignment=1))],
            [Paragraph(duree_fmt, S("tdur", fontSize=16, fontName="Helvetica-Bold",
                                    textColor=duree_col, alignment=1))],
            [Paragraph(f"{dm or '—'} minutes", S("tmin", fontSize=8, textColor=SLATE, alignment=1))],
        ]]),
        Paragraph("→", S("arr2", fontSize=16, textColor=SLATE, alignment=1)),
        Table([[
            [Paragraph("FIN D'INTERVENTION", S("tl3", fontSize=8, fontName="Helvetica-Bold", textColor=SLATE, alignment=2))],
            [Paragraph(fmt_date(inter.get("date_fin")), S("td3", fontSize=13, fontName="Helvetica-Bold", textColor=TEXT, alignment=2))],
            [Paragraph(inter.get("heure_fin") or "—", S("th3", fontSize=12, fontName="Helvetica-Bold", textColor=GREEN, alignment=2))],
        ]]),
    ]], colWidths=[5.5*cm, 1.2*cm, 4*cm, 1.2*cm, 5.5*cm])
    timeline.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1), LIGHT_GRAY),
        ("BOX",(0,0),(-1,-1),0.5, MID_GRAY),
        ("TOPPADDING",(0,0),(-1,-1),12),("BOTTOMPADDING",(0,0),(-1,-1),12),
        ("LEFTPADDING",(0,0),(-1,-1),12),("RIGHTPADDING",(0,0),(-1,-1),12),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
    ]))
    elements.append(timeline)
    elements.append(Spacer(1, 0.4*cm))

    # ── INFOS DÉTAILLÉES ────────────────────────────────────────
    def section(title):
        elements.append(Spacer(1, 0.2*cm))
        t = Table([[Paragraph(title, section_s)]], colWidths=[W])
        t.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1), DARK_BLUE),
            ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),
            ("LEFTPADDING",(0,0),(-1,-1),10),
            ("TEXTCOLOR",(0,0),(-1,-1), colors.white),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 0.15*cm))

    def row2(label, value):
        return [
            Paragraph(label, label_s),
            Paragraph(str(value) if value else "—", value_sm)
        ]

    section("INFORMATIONS GÉNÉRALES")
    info_rows = [
        row2("TYPE D'INTERVENTION", inter.get("type_intervention", "—")),
        row2("TECHNICIEN",          inter.get("technicien") or "—"),
        row2("MOIS / ANNÉE",        f"{(inter.get('mois') or '').capitalize()} {inter.get('annee') or ''}"),
    ]
    if prochaine_str:
        info_rows.append(row2("PROCHAINE INTERVENTION", prochaine_str))
    info_tbl = Table(info_rows, colWidths=[4.5*cm, 12.9*cm])
    info_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(0,-1), LIGHT_GRAY),
        ("TOPPADDING",(0,0),(-1,-1),7),("BOTTOMPADDING",(0,0),(-1,-1),7),
        ("LEFTPADDING",(0,0),(-1,-1),10),
        ("LINEBELOW",(0,0),(-1,-2),0.4, MID_GRAY),
        ("VALIGN",(0,0),(-1,-1),"TOP"),
    ]))
    elements.append(info_tbl)

    # ── TRAVAUX ──────────────────────────────────────────────────
    if inter.get("travaux"):
        section("TRAVAUX EFFECTUÉS")
        elements.append(Paragraph(inter["travaux"], body_s))

    # ── NOTES ────────────────────────────────────────────────────
    if inter.get("notes"):
        section("NOTES INTERNES")
        elements.append(Paragraph(inter["notes"], body_s))

    # ── PIÈCES JOINTES ───────────────────────────────────────────
    if pjs:
        section("PIÈCES JOINTES")
        pj_rows = [[
            Paragraph("NOM DU FICHIER", label_s),
            Paragraph("DATE", label_s),
        ]]
        for p in pjs:
            dt = (p.get("created_at") or "")[:16].replace("T", " ")
            pj_rows.append([
                Paragraph(p.get("nom_original") or p.get("nom_stockage", "—"), value_sm),
                Paragraph(dt, small_s),
            ])
        pj_tbl = Table(pj_rows, colWidths=[13*cm, 4.4*cm])
        pj_tbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0), DARK_BLUE),
            ("TEXTCOLOR",(0,0),(-1,0), colors.white),
            ("BACKGROUND",(0,1),(-1,-1), LIGHT_GRAY),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white, LIGHT_GRAY]),
            ("TOPPADDING",(0,0),(-1,-1),7),("BOTTOMPADDING",(0,0),(-1,-1),7),
            ("LEFTPADDING",(0,0),(-1,-1),10),
            ("LINEBELOW",(0,0),(-1,-2),0.4, MID_GRAY),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ]))
        elements.append(pj_tbl)

    # ── COMMENTAIRES ─────────────────────────────────────────────
    if comments:
        section("COMMENTAIRES")
        for c in comments:
            dt = (c.get("created_at") or "")[:16].replace("T", " ")
            author = c.get("username") or "—"
            row_c = Table([[
                Paragraph(f"<b>{author}</b>", S("ca", fontSize=9, fontName="Helvetica-Bold", textColor=PRIMARY)),
                Paragraph(dt, small_s),
            ]], colWidths=[10*cm, 7.4*cm])
            row_c.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,-1), BLUE_BG),
                ("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),
                ("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),10),
                ("ALIGN",(1,0),(1,-1),"RIGHT"),
            ]))
            elements.append(row_c)
            elements.append(Paragraph(c.get("contenu", ""), body_s))
            elements.append(Spacer(1, 0.2*cm))

    # ── FOOTER ───────────────────────────────────────────────────
    elements.append(Spacer(1, 0.6*cm))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=MID_GRAY))
    elements.append(Spacer(1, 0.2*cm))
    elements.append(Paragraph(
        f"Suivi des Interventions  ·  Rapport généré le {date.today().strftime('%d/%m/%Y')} "
        f"par {user['username']}  ·  Intervention N°{inter['id']}",
        footer_s))

    doc.build(elements)
    buf.seek(0)
    fname = f"intervention_{intervention_id}_{date.today().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


# ============================================================
# BACKUP BASE DE DONNÉES
# ============================================================

@app.post("/api/admin/backup")
def backup_database(request: Request):
    """Crée une sauvegarde de la base SQLite."""
    require_admin(request)
    user = get_current_user(request)
    backup_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "backups")
    os.makedirs(backup_dir, exist_ok=True)
    fname = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
    dest = os.path.join(backup_dir, fname)
    shutil.copy2(DB_PATH, dest)
    size = os.path.getsize(dest)
    conn = get_db()
    audit(conn, user, "backup", detail=fname)
    conn.commit()
    conn.close()
    return {"filename": fname, "size_bytes": size, "created_at": datetime.now().isoformat()}


@app.get("/api/admin/backup/download")
def telecharger_backup(request: Request):
    """Télécharge la DB en live."""
    require_admin(request)
    if not os.path.exists(DB_PATH):
        raise HTTPException(status_code=404, detail="Base de données non trouvée")
    fname = f"backup_{date.today().strftime('%Y%m%d')}.db"
    return FileResponse(DB_PATH, filename=fname, media_type="application/octet-stream")


@app.get("/api/admin/backups")
def liste_backups(request: Request):
    """Liste les sauvegardes disponibles."""
    require_admin(request)
    backup_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "backups")
    if not os.path.exists(backup_dir):
        return []
    files = []
    for f in sorted(os.listdir(backup_dir), reverse=True):
        if f.endswith(".db"):
            fp = os.path.join(backup_dir, f)
            files.append({"filename": f, "size_bytes": os.path.getsize(fp),
                          "created_at": datetime.fromtimestamp(os.path.getmtime(fp)).isoformat()})
    return files[:20]


# --- Duplication ---
@app.post("/api/interventions/{intervention_id}/dupliquer", status_code=201)
def dupliquer_intervention(intervention_id: int):
    """Duplique une intervention existante (sans les dates)."""
    conn = get_db()
    orig = conn.execute("SELECT * FROM interventions WHERE id=?", (intervention_id,)).fetchone()
    if not orig:
        conn.close()
        raise HTTPException(status_code=404, detail="Intervention non trouvée")
    cur = conn.execute("""
        INSERT INTO interventions
        (prestataire, type_intervention, mois, annee, date_debut, heure_debut,
         date_fin, heure_fin, duree_minutes, site, travaux, prochaine_intervention, notes)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (orig["prestataire"], orig["type_intervention"], orig["mois"], orig["annee"],
          None, None, None, None, None,
          orig["site"], orig["travaux"], orig["prochaine_intervention"],
          f"[Copie de #{intervention_id}] {orig['notes'] or ''}".strip()))
    new_id = cur.lastrowid

    # Historique : création par duplication
    enregistrer_historique(conn, new_id, "creation",
                           champ="source",
                           nouvelle=f"Copie de #{intervention_id}")

    conn.commit()
    row = conn.execute("SELECT * FROM interventions WHERE id=?", (new_id,)).fetchone()
    conn.close()
    return row_to_dict(row)


# ============================================================
# STATISTIQUES
# ============================================================

@app.get("/api/stats")
def statistiques(
    prestataire: Optional[str] = Query(None),
    type_intervention: Optional[str] = Query(None),
    mois: Optional[str] = Query(None),
    annee: Optional[int] = Query(None),
    site: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
):
    conn = get_db()
    where, params = build_where(prestataire, type_intervention, mois, annee, site, search)

    total = conn.execute(f"SELECT COUNT(*) FROM interventions {where}", params).fetchone()[0]

    # Par prestataire
    rows = conn.execute(
        f"SELECT prestataire, COUNT(*) as cnt FROM interventions {where} GROUP BY prestataire ORDER BY cnt DESC",
        params
    ).fetchall()
    par_prestataire = {r["prestataire"]: r["cnt"] for r in rows}

    # Par type
    rows = conn.execute(
        f"SELECT type_intervention, COUNT(*) as cnt FROM interventions {where} GROUP BY type_intervention",
        params
    ).fetchall()
    par_type = {}
    for r in rows:
        t = r["type_intervention"] or "Inconnu"
        tl = t.lower()
        if "plan" in tl: t = "Planifiée"
        elif "dep" in tl or "dép" in tl: t = "Dépannage"
        par_type[t] = par_type.get(t, 0) + r["cnt"]

    # Par mois (ordre calendaire)
    ordre = ["janvier","février","mars","avril","mai","juin",
             "juillet","août","septembre","octobre","novembre","décembre"]
    rows = conn.execute(
        f"SELECT mois, COUNT(*) as cnt FROM interventions {where} GROUP BY mois", params
    ).fetchall()
    raw = {(r["mois"] or "").lower(): r["cnt"] for r in rows}
    par_mois = {}
    for m in ordre:
        if m in raw: par_mois[m.capitalize()] = raw[m]
    for m, cnt in raw.items():
        if m not in ordre: par_mois[m.capitalize() or "Inconnu"] = cnt

    # Par site (top 10)
    rows = conn.execute(
        f"SELECT site, COUNT(*) as cnt FROM interventions {where} GROUP BY site ORDER BY cnt DESC LIMIT 10",
        params
    ).fetchall()
    par_site = {(r["site"] or "Inconnu"): r["cnt"] for r in rows}

    # Par statut
    rows = conn.execute(
        f"SELECT statut, COUNT(*) as cnt FROM interventions {where} GROUP BY statut ORDER BY cnt DESC",
        params
    ).fetchall()
    par_statut = {(r["statut"] or "Inconnu"): r["cnt"] for r in rows}

    # Durée moyenne
    row = conn.execute(
        f"SELECT AVG(duree_minutes) FROM interventions {where} AND duree_minutes IS NOT NULL AND duree_minutes > 0",
        params
    ).fetchone()
    duree_moyenne = round(row[0], 1) if row[0] else 0

    # Durée max
    row2 = conn.execute(
        f"SELECT MAX(duree_minutes) FROM interventions {where} AND duree_minutes IS NOT NULL",
        params
    ).fetchone()
    duree_max = row2[0] or 0

    # Durée totale
    row3 = conn.execute(
        f"SELECT SUM(duree_minutes) FROM interventions {where} AND duree_minutes IS NOT NULL",
        params
    ).fetchone()
    duree_totale = row3[0] or 0

    # Interventions longues (> 8h = 480 min)
    seuil = 480
    nb_longues = conn.execute(
        f"SELECT COUNT(*) FROM interventions {where} AND duree_minutes > ?",
        params + [seuil]
    ).fetchone()[0]

    # KPI
    nb_prestataires = conn.execute(
        f"SELECT COUNT(DISTINCT prestataire) FROM interventions {where}", params
    ).fetchone()[0]
    nb_sites = conn.execute(
        f"SELECT COUNT(DISTINCT site) FROM interventions {where}", params
    ).fetchone()[0]

    # Évolution dans le temps (par mois + année pour courbe)
    rows = conn.execute(
        f"""SELECT annee, mois, COUNT(*) as cnt FROM interventions {where}
            GROUP BY annee, mois ORDER BY annee, mois""",
        params
    ).fetchall()
    evolution = [{"annee": r["annee"], "mois": r["mois"], "cnt": r["cnt"]} for r in rows]

    # Années disponibles
    rows = conn.execute(
        "SELECT DISTINCT annee FROM interventions WHERE annee IS NOT NULL ORDER BY annee DESC"
    ).fetchall()
    annees_dispo = [r["annee"] for r in rows]

    # Prochaines interventions
    today = date.today().isoformat()
    rows = conn.execute("""
        SELECT *, (julianday(prochaine_intervention) - julianday(?)) as jours_restants
        FROM interventions
        WHERE prochaine_intervention IS NOT NULL AND prochaine_intervention != ''
        AND prochaine_intervention >= ?
        ORDER BY prochaine_intervention ASC LIMIT 5
    """, (today, today)).fetchall()
    prochaines = [row_to_dict(r) for r in rows]

    # Interventions longues détail
    rows = conn.execute(
        f"SELECT * FROM interventions {where} AND duree_minutes > ? ORDER BY duree_minutes DESC LIMIT 5",
        params + [seuil]
    ).fetchall()
    interventions_longues = [row_to_dict(r) for r in rows]

    conn.close()

    return {
        "total": total,
        "par_prestataire": par_prestataire,
        "par_type": par_type,
        "par_mois": par_mois,
        "par_site": par_site,
        "par_statut": par_statut,
        "duree_moyenne_minutes": duree_moyenne,
        "duree_max_minutes": duree_max,
        "duree_totale_minutes": duree_totale,
        "nb_prestataires": nb_prestataires,
        "nb_sites": nb_sites,
        "nb_interventions_longues": nb_longues,
        "evolution": evolution,
        "annees_disponibles": annees_dispo,
        "prochaines_interventions": prochaines,
        "interventions_longues": interventions_longues,
    }


# ============================================================
# STATS PRESTATAIRES (pour bubble chart)
# ============================================================

@app.get("/api/stats/prestataires")
def stats_prestataires():
    """
    Retourne des statistiques détaillées par prestataire pour le bubble chart.
    Inclut : nb_interventions, duree_moyenne, duree_totale, nb_sites.
    """
    conn = get_db()
    rows = conn.execute("""
        SELECT
            prestataire,
            COUNT(*) as nb_interventions,
            ROUND(AVG(CASE WHEN duree_minutes > 0 THEN duree_minutes END), 1) as duree_moyenne,
            SUM(CASE WHEN duree_minutes > 0 THEN duree_minutes ELSE 0 END) as duree_totale,
            COUNT(DISTINCT site) as nb_sites
        FROM interventions
        WHERE prestataire IS NOT NULL AND prestataire != ''
        GROUP BY prestataire
        ORDER BY nb_interventions DESC
    """).fetchall()
    conn.close()
    return [
        {
            "prestataire": r["prestataire"],
            "nb_interventions": r["nb_interventions"],
            "duree_moyenne": r["duree_moyenne"] or 0,
            "duree_totale": r["duree_totale"] or 0,
            "nb_sites": r["nb_sites"],
        }
        for r in rows
    ]


# ============================================================
# EXPORT CSV
# ============================================================

@app.get("/api/export/csv")
def export_csv(request: Request):
    """Export CSV lisible par Excel (UTF-8 BOM, séparateur ;)."""
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Non authentifié")
    conn = get_db()
    rows = conn.execute("SELECT * FROM interventions ORDER BY date_debut DESC").fetchall()
    conn.close()

    output = io.BytesIO()
    output.write(b'\xef\xbb\xbf')  # BOM UTF-8 pour Excel

    text = io.StringIO()
    writer = csv.writer(text, delimiter=";", quoting=csv.QUOTE_MINIMAL)

    writer.writerow([
        "ID", "Prestataire", "Type intervention", "Mois", "Annee",
        "Date debut", "Heure debut", "Date fin", "Heure fin",
        "Duree (HH:MM)", "Site", "Travaux", "Prochaine intervention", "Notes", "Cree le"
    ])

    for r in rows:
        def fmt_date(d):
            if not d: return ""
            try:
                p = d.split("-")
                return f"{p[2]}/{p[1]}/{p[0]}" if len(p) == 3 else d
            except: return d or ""

        dm = r["duree_minutes"]
        duree_fmt = f"{dm//60:02d}:{dm%60:02d}" if dm is not None and dm >= 0 else ""

        writer.writerow([
            r["id"], r["prestataire"] or "", r["type_intervention"] or "",
            r["mois"] or "", r["annee"] or "",
            fmt_date(r["date_debut"]), r["heure_debut"] or "",
            fmt_date(r["date_fin"]), r["heure_fin"] or "",
            duree_fmt, r["site"] or "", r["travaux"] or "",
            fmt_date(r["prochaine_intervention"]),
            r["notes"] or "", r["created_at"] or "",
        ])

    output.write(text.getvalue().encode("utf-8"))
    output.seek(0)
    fname = f"interventions_{date.today().strftime('%Y%m%d')}.csv"
    return StreamingResponse(
        output,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={fname}", "Cache-Control": "no-cache"}
    )


# ============================================================
# EXPORT EXCEL (.xlsx)
# ============================================================

@app.get("/api/export/excel")
def export_excel(request: Request):
    """Export Excel .xlsx avec mise en forme."""
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Non authentifié")
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl non installé")

    conn = get_db()
    rows = conn.execute("SELECT * FROM interventions ORDER BY date_debut DESC").fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Interventions"

    # Styles
    header_font    = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", fgColor="1E3A8A")
    header_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin           = Side(style="thin", color="CBD5E1")
    border         = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill       = PatternFill("solid", fgColor="EFF6FF")
    plan_fill      = PatternFill("solid", fgColor="D1FAE5")
    dep_fill       = PatternFill("solid", fgColor="FEE2E2")
    center_align   = Alignment(horizontal="center", vertical="center")
    wrap_align     = Alignment(wrap_text=True, vertical="top")

    headers = [
        "N°", "Prestataire", "Type", "Mois", "Année",
        "Date début", "Heure début", "Date fin", "Heure fin",
        "Durée (HH:MM)", "Site", "Travaux", "Prochaine intervention", "Notes"
    ]
    col_widths = [6, 16, 14, 10, 7, 13, 11, 13, 11, 13, 22, 50, 18, 30]

    ws.row_dimensions[1].height = 28
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    def fmt_date(d):
        if not d: return ""
        try:
            p = d.split("-")
            return f"{p[2]}/{p[1]}/{p[0]}" if len(p) == 3 else d
        except: return d or ""

    for row_idx, r in enumerate(rows, 2):
        dm = r["duree_minutes"]
        duree_fmt = f"{dm//60:02d}:{dm%60:02d}" if dm is not None and dm >= 0 else ""
        type_inter = r["type_intervention"] or ""
        is_plan = "plan" in type_inter.lower()

        vals = [
            r["id"], r["prestataire"] or "", type_inter,
            r["mois"] or "", r["annee"] or "",
            fmt_date(r["date_debut"]), r["heure_debut"] or "",
            fmt_date(r["date_fin"]), r["heure_fin"] or "",
            duree_fmt, r["site"] or "", r["travaux"] or "",
            fmt_date(r["prochaine_intervention"]), r["notes"] or "",
        ]

        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            if col_idx == 3:
                cell.fill = plan_fill if is_plan else dep_fill
                cell.alignment = center_align
                cell.font = Font(bold=True, color="065F46" if is_plan else "991B1B")
            if col_idx in (12, 14):
                cell.alignment = wrap_align
            if col_idx == 10 and dm and dm > 480:
                cell.font = Font(bold=True, color="DC2626")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Onglet Stats résumé
    ws2 = wb.create_sheet("Résumé")
    ws2["A1"] = "Statistiques"
    ws2["A1"].font = Font(bold=True, size=14, color="1E3A8A")
    ws2["A3"] = "Total interventions"
    ws2["B3"] = len(rows)
    ws2["A4"] = "Fichier généré le"
    ws2["B4"] = date.today().strftime("%d/%m/%Y")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    fname = f"interventions_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


# ============================================================
# IMPORT CSV DEPUIS L'INTERFACE
# ============================================================

@app.post("/api/import/csv")
async def import_csv(file: UploadFile = File(...), request: Request = None):
    """Importe des interventions depuis un fichier CSV uploadé."""
    user = get_current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Non authentifié")
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Fichier CSV requis (.csv)")

    content = await file.read()

    # Détection encodage
    for enc in ["utf-8-sig", "utf-8", "latin-1", "cp1252"]:
        try:
            text = content.decode(enc)
            break
        except Exception:
            text = None

    if not text:
        raise HTTPException(status_code=400, detail="Encodage du fichier non reconnu")

    MOIS_MAP = {
        "janvier":"janvier","fevrier":"février","février":"février",
        "mars":"mars","avril":"avril","mai":"mai","juin":"juin",
        "juillet":"juillet","aout":"août","août":"août",
        "septembre":"septembre","octobre":"octobre",
        "novembre":"novembre","decembre":"décembre","décembre":"décembre",
    }
    PREST_MAP = {
        "dsi":"DSI","cogitech":"COGITECH","manage engine":"Manage Engine",
        "cbi":"CBI","gbss":"GBSS","awale":"AWALE","prosuma":"PROSUMA",
        "orange":"ORANGE","mtn":"MTN","aric":"ARIC",
    }

    import re
    def parse_date(s):
        if not s: return None
        s = re.sub(r'\s+','',s.strip())
        s = re.sub(r'/(\d{5})$', lambda m: '/'+m.group(1)[1:], s)
        for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"]:
            try: return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except: pass
        return None

    def parse_heure(s):
        if not s: return None
        h = s.strip()
        return h if re.match(r'^\d{2}:\d{2}$', h) else None

    reader = csv.reader(text.splitlines(), delimiter=";")
    lignes = list(reader)

    importees = 0
    erreurs = 0
    ignorees = 0

    conn = get_db()

    for i, row in enumerate(lignes):
        if i < 2: continue  # Ignorer titre + en-tête
        if len(row) < 2: ignorees += 1; continue
        prest_raw = row[1].strip() if len(row) > 1 else ""
        if not prest_raw: ignorees += 1; continue

        try:
            prestataire = PREST_MAP.get(prest_raw.lower(), prest_raw.strip())
            type_raw    = row[2].strip() if len(row) > 2 else ""
            tl = type_raw.lower()
            type_inter  = "Planifiée" if "plan" in tl else ("Dépannage" if "dep" in tl or "dép" in tl else type_raw.capitalize())
            mois_raw    = row[3].strip() if len(row) > 3 else ""
            mois        = MOIS_MAP.get(mois_raw.lower().replace("é","e").replace("û","u"), mois_raw.lower() or None)
            annee_raw   = row[4].strip() if len(row) > 4 else ""
            annee       = int(annee_raw) if annee_raw.isdigit() else None
            date_debut  = parse_date(row[5].strip() if len(row) > 5 else "")
            heure_debut = parse_heure(row[6].strip() if len(row) > 6 else "")
            date_fin    = parse_date(row[7].strip() if len(row) > 7 else "")
            heure_fin   = parse_heure(row[8].strip() if len(row) > 8 else "")
            site        = row[10].strip() if len(row) > 10 else ""
            travaux     = row[11].strip() if len(row) > 11 else ""
            prochaine   = parse_date(row[12].strip() if len(row) > 12 else "")

            duree = calcul_duree(date_debut, heure_debut, date_fin, heure_fin)
            if duree is None:
                d_raw = row[9].strip() if len(row) > 9 else ""
                m2 = re.match(r'^(\d+):(\d{2})$', d_raw)
                if m2: duree = int(m2.group(1))*60 + int(m2.group(2))

            if date_debut and not mois:
                try:
                    d = datetime.strptime(date_debut, "%Y-%m-%d")
                    noms = ["janvier","février","mars","avril","mai","juin","juillet","août","septembre","octobre","novembre","décembre"]
                    mois = noms[d.month-1]
                    if not annee: annee = d.year
                except: pass

            cur = conn.execute("""
                INSERT INTO interventions
                (prestataire, type_intervention, mois, annee, date_debut, heure_debut,
                 date_fin, heure_fin, duree_minutes, site, travaux, prochaine_intervention)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
            """, (prestataire, type_inter, mois, annee, date_debut, heure_debut,
                  date_fin, heure_fin, duree, site or None, travaux or None, prochaine))

            # Historique import
            enregistrer_historique(conn, cur.lastrowid, "creation",
                                   champ="source", nouvelle="import_csv")
            importees += 1
        except Exception:
            erreurs += 1

    conn.commit()
    conn.close()
    return {"importees": importees, "erreurs": erreurs, "ignorees": ignorees}


# ============================================================
# UTILITAIRES
# ============================================================

@app.get("/api/annees")
def liste_annees():
    """Retourne les années disponibles dans la DB."""
    conn = get_db()
    rows = conn.execute(
        "SELECT DISTINCT annee FROM interventions WHERE annee IS NOT NULL ORDER BY annee DESC"
    ).fetchall()
    conn.close()
    return [r["annee"] for r in rows]


# (routes HTML et fichiers statiques déjà déclarés plus haut)
